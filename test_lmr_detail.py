"""
L側X座標の計算過程を詳細に出力するテストスクリプト
"""

import math
import openpyxl

def detailed_calculation():
    """L側X座標（Q列）の計算を詳細に出力"""
    
    print("=" * 80)
    print("L側X座標（Q列）の計算過程詳細")
    print("=" * 80)
    
    # Excelファイルから実際の値を読み込む
    excel_path = "data/江川トンネル進行表.xlsx"
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb.active
    
    # 1245行目と974行目の値を取得
    print("\n【1. Excelから値を読み込み】")
    print("-" * 40)
    
    # I列の値（トンネル坑口からの距離）
    I_974 = sheet['I974'].value
    I_1245 = sheet['I1245'].value
    print(f"I974 (基準点の距離): {I_974}")
    print(f"I1245 (計算対象の距離): {I_1245}")
    
    # W列の値（角度）
    W_1245 = sheet['W1245'].value
    print(f"W1245 (角度): {W_1245}°")
    
    # Q列の値（基準座標）
    Q_974 = sheet['Q974'].value
    print(f"Q974 (基準点のX座標): {Q_974}")
    
    # Q1245の実際の値（期待値）
    Q_1245_expected = sheet['Q1245'].value
    print(f"Q1245 (期待される計算結果): {Q_1245_expected}")
    
    wb.close()
    
    # Excel数式の再現
    # =ROUND((-($I1245-$I$974)*1000*COS((90-$W1245)*PI()/180)+Q$974)/1000,3)
    
    print("\n【2. 計算式の分解】")
    print("-" * 40)
    print("Excel数式: =ROUND((-($I1245-$I$974)*1000*COS((90-$W1245)*PI()/180)+Q$974)/1000,3)")
    
    print("\n【3. ステップごとの計算】")
    print("-" * 40)
    
    # Step 1: 距離の差分
    distance_diff = I_1245 - I_974
    print(f"Step 1: 距離の差分 = I1245 - I974")
    print(f"        = {I_1245} - {I_974}")
    print(f"        = {distance_diff} m")
    
    # Step 2: メートルをミリメートルに変換
    distance_diff_mm = distance_diff * 1000
    print(f"\nStep 2: ミリメートル変換 = {distance_diff} * 1000")
    print(f"        = {distance_diff_mm} mm")
    
    # Step 3: 角度の計算
    angle_calc = 90 - W_1245
    print(f"\nStep 3: 角度の計算 = 90 - W1245")
    print(f"        = 90 - {W_1245}")
    print(f"        = {angle_calc}°")
    
    # Step 4: ラジアンに変換
    angle_rad = angle_calc * math.pi / 180
    print(f"\nStep 4: ラジアン変換 = {angle_calc} * π / 180")
    print(f"        = {angle_calc} * {math.pi} / 180")
    print(f"        = {angle_rad} rad")
    
    # Step 5: コサイン計算
    cos_value = math.cos(angle_rad)
    print(f"\nStep 5: コサイン計算 = cos({angle_rad})")
    print(f"        = {cos_value}")
    
    # Step 6: 距離とコサインの積
    distance_cos_product = distance_diff_mm * cos_value
    print(f"\nStep 6: 距離×コサイン = {distance_diff_mm} * {cos_value}")
    print(f"        = {distance_cos_product}")
    
    # Step 7: 符号を反転
    negative_product = -distance_cos_product
    print(f"\nStep 7: 符号反転 = -({distance_cos_product})")
    print(f"        = {negative_product}")
    
    # Step 8: 基準座標を加算
    with_reference = negative_product + Q_974
    print(f"\nStep 8: 基準座標を加算 = {negative_product} + {Q_974}")
    print(f"        = {with_reference}")
    
    # Step 9: ミリメートルからメートルに変換
    result_before_round = with_reference / 1000
    print(f"\nStep 9: メートル変換 = {with_reference} / 1000")
    print(f"        = {result_before_round}")
    
    # Step 10: 小数第3位で丸める
    result_final = round(result_before_round, 3)
    print(f"\nStep 10: 丸め処理 = round({result_before_round}, 3)")
    print(f"         = {result_final}")
    
    print("\n【4. 結果の比較】")
    print("-" * 40)
    print(f"計算結果: {result_final}")
    print(f"期待値:   {Q_1245_expected}")
    difference = abs(result_final - Q_1245_expected) if Q_1245_expected else 0
    print(f"差:       {difference}")
    
    if difference < 0.001:
        print("✅ 計算結果が期待値と一致しています！")
    else:
        print("❌ 計算結果が期待値と異なります")
        
        # デバッグ情報
        print("\n【5. デバッグ情報】")
        print("-" * 40)
        
        # 別の計算方法で検証
        print("\n別の計算方法での検証:")
        
        # Q974がすでにメートル単位の可能性を検証
        print(f"\n仮説1: Q974がすでにメートル単位の場合")
        alt_calc1 = round((-distance_diff * 1000 * cos_value + Q_974 * 1000) / 1000, 3)
        print(f"  計算結果: {alt_calc1}")
        
        # 基準座標の単位が異なる可能性
        print(f"\n仮説2: 基準座標の処理が異なる場合")
        alt_calc2 = round(-distance_diff * cos_value + Q_974 / 1000, 3)
        print(f"  計算結果: {alt_calc2}")
        
        # 直接計算
        print(f"\n仮説3: シンプルな計算")
        alt_calc3 = round(Q_974 - distance_diff * cos_value, 3)
        print(f"  計算結果: {alt_calc3}")
    
    return result_final, Q_1245_expected

def check_formula_interpretation():
    """数式の解釈を確認"""
    print("\n" + "=" * 80)
    print("数式の解釈確認")
    print("=" * 80)
    
    excel_path = "data/江川トンネル進行表.xlsx"
    
    # 数式を直接読み取る（data_only=False）
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    sheet = wb.active
    
    # Q1245の数式を取得
    formula_Q1245 = sheet['Q1245'].value
    print(f"\nQ1245の実際の数式: {formula_Q1245}")
    
    # 他の関連する数式も確認
    print(f"\nR1245の数式: {sheet['R1245'].value}")
    print(f"S1245の数式: {sheet['S1245'].value}")
    
    wb.close()
    
    # 値を再度確認
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb.active
    
    print("\n【関連する値の確認】")
    print(f"I974:  {sheet['I974'].value}")
    print(f"I1245: {sheet['I1245'].value}")
    print(f"W1245: {sheet['W1245'].value}")
    print(f"Q974:  {sheet['Q974'].value}")
    print(f"Q1245: {sheet['Q1245'].value}")
    
    wb.close()

if __name__ == "__main__":
    print("L側X座標の計算過程を詳細に解析します...\n")
    
    # 詳細計算
    calculated, expected = detailed_calculation()
    
    # 数式の解釈確認
    check_formula_interpretation()
    
    print("\n" + "=" * 80)
    print("解析完了")
    print("=" * 80)